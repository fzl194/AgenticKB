from __future__ import annotations

from datetime import date, datetime, time
import math
from pathlib import Path
import re
from typing import Any

from knowledge_mining.mining.ingestion.errors import PreprocessingError
from knowledge_mining.mining.ingestion.excel_models import (
    ExcelWarning,
    NormalizedSheet,
    NormalizedWorkbook,
)


def read_excel_workbook(path: Path) -> NormalizedWorkbook:
    path = Path(path)
    suffix = path.suffix.lower()
    try:
        if suffix == ".xlsx":
            return _read_xlsx(path)
        if suffix == ".xls":
            return _read_xls(path)
        raise PreprocessingError(
            "excel_unsupported_format", f"Unsupported Excel format: {suffix}"
        )
    except PreprocessingError:
        raise
    except ImportError as exc:
        raise PreprocessingError(
            "excel_dependency_missing", "Excel parser dependency is not installed"
        ) from exc
    except Exception as exc:
        message = str(exc).lower()
        code = (
            "excel_password_protected"
            if "password" in message or "encrypted" in message
            else "excel_corrupt_file"
        )
        raise PreprocessingError(
            code, f"Unable to read Excel workbook: {path.name}"
        ) from exc


def _read_xlsx(path: Path) -> NormalizedWorkbook:
    from openpyxl import load_workbook

    values_workbook = load_workbook(
        path,
        data_only=True,
        read_only=False,
        keep_links=False,
    )
    try:
        formula_workbook = load_workbook(
            path,
            data_only=False,
            read_only=False,
            keep_links=False,
        )
    except Exception:
        values_workbook.close()
        raise

    sheets: list[NormalizedSheet] = []
    warnings: list[ExcelWarning] = []
    try:
        for sheet_name in values_workbook.sheetnames:
            try:
                value_sheet = values_workbook[sheet_name]
                formula_sheet = formula_workbook[sheet_name]
                sheets.append(_materialize_xlsx_sheet(value_sheet, formula_sheet))
            except Exception:
                warnings.append(
                    ExcelWarning(
                        code="excel_sheet_parse_failed",
                        message="Unable to read Excel worksheet",
                        sheet_name=sheet_name,
                    )
                )
    finally:
        formula_workbook.close()
        values_workbook.close()

    return NormalizedWorkbook(
        source_format="xlsx",
        sheets=tuple(sheets),
        warnings=tuple(warnings),
    )


def _materialize_xlsx_sheet(value_sheet: Any, formula_sheet: Any) -> NormalizedSheet:
    max_row = max(value_sheet.max_row, formula_sheet.max_row)
    max_column = max(value_sheet.max_column, formula_sheet.max_column)
    merged_values: dict[tuple[int, int], Any] = {}

    for merged_range in value_sheet.merged_cells.ranges:
        top_left = value_sheet.cell(merged_range.min_row, merged_range.min_col).value
        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            for column_index in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row_index, column_index)] = top_left

    rows: list[tuple[str, ...]] = []
    sheet_warnings: list[ExcelWarning] = []
    for row_index in range(1, max_row + 1):
        values: list[str] = []
        for column_index in range(1, max_column + 1):
            cached_cell = value_sheet.cell(row_index, column_index)
            formula_cell = formula_sheet.cell(row_index, column_index)
            coordinate = cached_cell.coordinate

            if formula_cell.data_type == "f" and cached_cell.value is None:
                sheet_warnings.append(
                    ExcelWarning(
                        code="excel_formula_cache_missing",
                        message="Formula has no saved calculation result",
                        sheet_name=value_sheet.title,
                        cell_range=coordinate,
                    )
                )

            raw_value = merged_values.get(
                (row_index, column_index), cached_cell.value
            )
            if cached_cell.data_type == "e":
                sheet_warnings.append(
                    ExcelWarning(
                        code="excel_cell_error",
                        message=f"Excel cell contains error {raw_value}",
                        sheet_name=value_sheet.title,
                        cell_range=coordinate,
                    )
                )
            values.append(_normalize_scalar(raw_value, cached_cell.number_format))
        rows.append(tuple(values))

    return NormalizedSheet(
        name=value_sheet.title,
        hidden=value_sheet.sheet_state != "visible",
        rows=tuple(rows),
        formula_warnings=tuple(sheet_warnings),
    )


def _read_xls(path: Path) -> NormalizedWorkbook:
    import xlrd

    workbook = xlrd.open_workbook(
        str(path),
        formatting_info=True,
        on_demand=True,
    )
    sheets: list[NormalizedSheet] = []
    warnings: list[ExcelWarning] = []
    try:
        for sheet_name in workbook.sheet_names():
            try:
                sheet = workbook.sheet_by_name(sheet_name)
                sheets.append(_materialize_xls_sheet(workbook, sheet, xlrd))
            except Exception:
                warnings.append(
                    ExcelWarning(
                        code="excel_sheet_parse_failed",
                        message="Unable to read Excel worksheet",
                        sheet_name=sheet_name,
                    )
                )
    finally:
        workbook.release_resources()

    return NormalizedWorkbook(
        source_format="xls",
        sheets=tuple(sheets),
        warnings=tuple(warnings),
    )


def _materialize_xls_sheet(workbook: Any, sheet: Any, xlrd: Any) -> NormalizedSheet:
    merged_sources: dict[tuple[int, int], tuple[int, int]] = {}
    for row_low, row_high, column_low, column_high in sheet.merged_cells:
        for row_index in range(row_low, row_high):
            for column_index in range(column_low, column_high):
                merged_sources[(row_index, column_index)] = (row_low, column_low)

    rows: list[tuple[str, ...]] = []
    sheet_warnings: list[ExcelWarning] = []
    for row_index in range(sheet.nrows):
        values: list[str] = []
        for column_index in range(sheet.ncols):
            source_row, source_column = merged_sources.get(
                (row_index, column_index), (row_index, column_index)
            )
            cell = sheet.cell(source_row, source_column)
            raw_value = cell.value
            number_format = _xls_number_format(workbook, cell)

            if cell.ctype == xlrd.XL_CELL_DATE:
                raw_value = xlrd.xldate_as_datetime(raw_value, workbook.datemode)
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                raw_value = bool(raw_value)
            elif cell.ctype == xlrd.XL_CELL_ERROR:
                raw_value = xlrd.error_text_from_code.get(
                    int(raw_value), f"#ERROR({int(raw_value)})"
                )
                sheet_warnings.append(
                    ExcelWarning(
                        code="excel_cell_error",
                        message=f"Excel cell contains error {raw_value}",
                        sheet_name=sheet.name,
                        cell_range=f"R{row_index + 1}C{column_index + 1}",
                    )
                )
            elif cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                raw_value = ""

            values.append(_normalize_scalar(raw_value, number_format))
        rows.append(tuple(values))

    return NormalizedSheet(
        name=sheet.name,
        hidden=bool(getattr(sheet, "visibility", 0)),
        rows=tuple(rows),
        formula_warnings=tuple(sheet_warnings),
    )


def _xls_number_format(workbook: Any, cell: Any) -> str:
    try:
        format_key = workbook.xf_list[cell.xf_index].format_key
        return workbook.format_map[format_key].format_str
    except (AttributeError, IndexError, KeyError, TypeError):
        return ""


def _normalize_scalar(value: Any, number_format: str = "") -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, (int, float)):
        numeric_value = float(value)
        if not math.isfinite(numeric_value):
            return str(value)
        if _is_percentage_format(number_format):
            decimals = _percentage_decimal_places(number_format)
            return f"{numeric_value * 100:.{decimals}f}%"
        return format(value, ".15g")
    return str(value).strip()


def _is_percentage_format(number_format: str) -> bool:
    return "%" in re.sub(r'"[^"]*"', "", number_format or "")


def _percentage_decimal_places(number_format: str) -> int:
    cleaned = re.sub(r'"[^"]*"', "", number_format or "")
    percent_section = cleaned.split("%", 1)[0]
    decimal_match = re.search(r"\.([0#]+)$", percent_section)
    return len(decimal_match.group(1)) if decimal_match else 0
