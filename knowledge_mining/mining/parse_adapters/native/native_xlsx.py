"""XLSX 原生解析适配器 + Normalizer（M3, SRS §C06/§C07, ADR-0003 D-028A；
2026-08-17 整改轮重写区域识别）.

理念：全部用工业级成熟库（openpyxl），自研代码只做"库输出 ->
BackendBlock"映射。

整改轮修正（用户指令）：
- **不再把整个 Sheet 无条件视为一张语义表**。识别优先级：
  ① Excel Table（``ws.tables``，命名表带真实表头语义）；
  ② 连续数据区域（空行/空列切分的连通块，origin=contiguous_region）；
  ③ 退化：单一区域覆盖全部使用范围时 origin=used_range（仍一张表，
     但策略可追溯）。区域数超上限截断并诊断。
- 公式（data_only=False）与展示值（data_only=True）双读保持不变；
- 合并区域展开保持；**隐藏行/列状态**进入 structure（可见，不静默）；
- **图表/图片当前不支持 -> 计数进 diagnostics**（不静默丢失）；
- 表格 Element.text 由统一 rendered_text 渲染（骨架保证）；
- cell 级 EvidenceSpan（native_ref=sheet+绝对 A1）经 ``_make_cell_spans``。

fingerprint：``native_xlsx@2.0.0#openpyxl-<ver>``（区域策略变更 -> 版本升）。
"""
from __future__ import annotations

import io
from importlib.metadata import version as _pkg_version
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from knowledge_mining.mining.contracts.parse_ir import (
    Container,
    EvidenceSpan,
)
from knowledge_mining.mining.contracts.parser_adapter import (
    BackendBlock,
    BackendParseArtifact,
    ParserAdapterError,
    ParserDescriptor,
    UnsupportedFormat,
)
from knowledge_mining.mining.parse_adapters.native._base import (
    BaseNativeNormalizer,
)

NATIVE_XLSX_PARSER_ID = "native_xlsx"
NATIVE_XLSX_VERSION = "2.0.0"
_OPENPYXL_VERSION = _pkg_version("openpyxl")
NATIVE_XLSX_FINGERPRINT = (
    f"{NATIVE_XLSX_PARSER_ID}@{NATIVE_XLSX_VERSION}"
    f"#openpyxl-{_OPENPYXL_VERSION}"
)

NATIVE_XLSX_MIMES = frozenset({
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
})

WORKBOOK_CONTAINER_ID = "c-workbook"

#: 区域识别上限（防碎片轰炸）：单 sheet 产出的最大语义表数。
_MAX_REGIONS_PER_SHEET = 64


class NativeXlsxParser:
    """DocumentParser 实现：openpyxl 双读包装（SRS §C06）."""

    def __init__(self) -> None:
        self.descriptor = ParserDescriptor(
            parser_id=NATIVE_XLSX_PARSER_ID,
            display_name="Native XLSX Parser (openpyxl)",
            version=NATIVE_XLSX_VERSION,
            supported_mimes=NATIVE_XLSX_MIMES,
            backend_kind="local",
            parser_fingerprint=NATIVE_XLSX_FINGERPRINT,
            capabilities=frozenset({
                "sheets", "excel_tables", "data_regions", "merges", "formulas",
            }),
        )

    def supports(self, mime: str) -> bool:
        return self.descriptor.supports(mime)

    def parse(self, data: bytes, *, mime: str) -> BackendParseArtifact:
        if not self.supports(mime):
            raise UnsupportedFormat(
                f"{NATIVE_XLSX_PARSER_ID} cannot parse mime {mime!r}"
            )
        try:
            wb_formula = load_workbook(io.BytesIO(data), data_only=False)
            wb_values = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as exc:  # 第三方异常不得穿越适配层（SRS §C06）
            raise ParserAdapterError(
                f"{NATIVE_XLSX_PARSER_ID}: openpyxl failed to open source: {exc}"
            ) from exc

        blocks: list[BackendBlock] = []
        warnings: list[str] = []
        try:
            for sheet_name in wb_formula.sheetnames:
                sheet_formula = wb_formula[sheet_name]
                sheet_values = wb_values[sheet_name]
                _emit_sheet_blocks(
                    sheet_formula, sheet_values, blocks, warnings
                )
        except Exception as exc:  # 中段损坏也归一（§C06，评审 MED）
            raise ParserAdapterError(
                f"{NATIVE_XLSX_PARSER_ID}: failed to walk sheets: {exc}"
            ) from exc

        return BackendParseArtifact(
            parser_id=NATIVE_XLSX_PARSER_ID,
            parser_version=NATIVE_XLSX_VERSION,
            mime=mime.lower(),
            blocks=tuple(blocks),
            warnings=tuple(warnings),
        )


# ---------------------------------------------------------------------------
# Parser 映射（模块级纯函数，便于单测）
# ---------------------------------------------------------------------------
#: 不可信声明几何上限（评审 HIGH-2）：单 sheet 遍历网格面积与合并区域
#: 面积的上限。超限的合并区域按未合并处理并计入 clamped_merges（§7.4
#: 可见性）；超限网格截断为 (MAX_GRID_EDGE, MAX_GRID_EDGE)。
_MAX_GRID_AREA = 2_000_000  # rows*cols
_MAX_GRID_EDGE = 10_000
_MAX_MERGE_AREA = 100_000


def _emit_sheet_blocks(
    sheet_formula: Worksheet,
    sheet_values: Worksheet,
    blocks: list[BackendBlock],
    warnings: list[str],
) -> None:
    """一个 sheet -> 0..N 个语义表块 + 诊断（图表/图片/空 sheet）."""
    _diagnose_unsupported(sheet_formula, warnings)

    regions = _detect_regions(sheet_formula)
    if not regions:
        return
    if len(regions) > _MAX_REGIONS_PER_SHEET:
        warnings.append(
            f"sheet {sheet_formula.title!r}: {len(regions)} regions found, "
            f"truncated to {_MAX_REGIONS_PER_SHEET}"
        )
        regions = regions[:_MAX_REGIONS_PER_SHEET]

    used = _used_bounds(sheet_formula)
    for region in regions:
        block = _region_block(sheet_formula, sheet_values, region, used)
        if block is not None:
            blocks.append(block)


def _diagnose_unsupported(sheet: Worksheet, warnings: list[str]) -> None:
    """图表/图片不支持 -> 计数诊断（不静默丢失，SRS §7.4）."""
    charts = len(getattr(sheet, "_charts", ()) or ())
    images = len(getattr(sheet, "_images", ()) or ())
    if charts:
        warnings.append(
            f"sheet {sheet.title!r}: {charts} chart(s) present; "
            "chart extraction not supported (diagnosed, not parsed)"
        )
    if images:
        warnings.append(
            f"sheet {sheet.title!r}: {images} image(s) present; "
            "image extraction not supported (diagnosed, not parsed)"
        )


def _used_bounds(sheet: Worksheet) -> tuple[int, int, int, int] | None:
    """sheet 使用范围（0-based row0/col0/row1/col1，含端点）；空 -> None."""
    max_row = sheet.max_row or 0
    max_col = sheet.max_column or 0
    if max_row <= 0 or max_col <= 0:
        return None
    return (0, 0, min(max_row, _MAX_GRID_EDGE) - 1,
            min(max_col, _MAX_GRID_EDGE) - 1)


def _detect_regions(sheet: Worksheet) -> list[dict[str, Any]]:
    """区域识别：Excel Table 优先 -> 连续数据区域 -> used_range 退化.

    返回 region dict：origin/row0/col0/row1/col1（0-based 含端点）/
    header_rows（Excel Table 真实表头行数，其余 0）。连续区域会被
    起点落在其内的合并区域扩展边界（合并覆盖格无值但有结构语义）。
    """
    regions: list[dict[str, Any]] = []
    # ① Excel Table（命名表）：真实表头语义。openpyxl 版本差异：
    # ``ws.tables`` 的值是 ref 字符串或 Table 对象（3.1.x 为字符串）。
    for name, table in getattr(sheet, "tables", {}).items():
        if isinstance(table, str):
            ref, header_rows = table, 1
        else:
            ref = table.ref
            header_rows = int(getattr(table, "headerRowCount", 1) or 0)
        min_col, min_row, max_col, max_row = _parse_ref(ref)
        regions.append({
            "origin": f"excel_table:{name}",
            "row0": min_row - 1, "col0": min_col - 1,
            "row1": max_row - 1, "col1": max_col - 1,
            "header_rows": header_rows,
            "ref": ref,
        })
    if regions:
        return regions

    # ② 连续数据区域：空行/空列切分连通块（只扫物化单元格，
    # 恶意稀疏声明不产生全网格遍历——评审 HIGH-2 的性能面）。
    coords = _nonempty_coords(sheet)
    if not coords:
        return []
    rows_present = sorted({r for r, _ in coords})
    row_bands = _bands_from_sorted(rows_present)

    regions = []
    for rb0, rb1 in row_bands:
        cols_present = sorted({c for r, c in coords if rb0 <= r <= rb1})
        for cb0, cb1 in _bands_from_sorted(cols_present):
            regions.append({
                "origin": "contiguous_region",
                "row0": rb0, "col0": cb0, "row1": rb1, "col1": cb1,
                "header_rows": 0,
                "ref": _a1_range(rb0, cb0, rb1, cb1),
            })

    # 合并区域扩展：起点落在区域内的 merge 把区域边界撑到 merge 端点
    # （值只在原点，连通块会低估结构范围）。
    for merged in sheet.merged_cells.ranges:
        mr0, mc0 = merged.min_row - 1, merged.min_col - 1
        mr1, mc1 = merged.max_row - 1, merged.max_col - 1
        for region in regions:
            if region["row0"] <= mr0 <= region["row1"] and \
                    region["col0"] <= mc0 <= region["col1"]:
                region["row1"] = max(region["row1"], mr1)
                region["col1"] = max(region["col1"], mc1)
                region["ref"] = _a1_range(
                    region["row0"], region["col0"],
                    region["row1"], region["col1"],
                )
                break

    # ③ 退化：单一区域覆盖全部使用范围 -> used_range（策略可追溯）
    if len(regions) == 1:
        r = regions[0]
        all_rows = rows_present + [
            m.max_row - 1 for m in sheet.merged_cells.ranges
        ]
        all_cols = (
            [c for _, c in coords]
            + [m.max_col - 1 for m in sheet.merged_cells.ranges]
        )
        if r["row0"] == min(all_rows) and r["col0"] == min(all_cols) \
                and r["row1"] == max(all_rows) and r["col1"] == max(all_cols):
            r["origin"] = "used_range"
    return regions


def _nonempty_coords(sheet: Worksheet) -> set[tuple[int, int]]:
    """非空单元格坐标集合（0-based）。

    只遍历**物化**单元格（``_cells``：读模式下恰好是文件中存在的格），
    避免对恶意稀疏声明（max_row/max_col 巨大）做全网格扫描。
    """
    coords: set[tuple[int, int]] = set()
    cells = getattr(sheet, "_cells", None)
    if cells is not None:
        for (row, col), cell in cells.items():
            value = getattr(cell, "value", None)
            if value is not None and str(value) != "":
                coords.add((row - 1, col - 1))
        return coords
    # 兜底：物化表缺失时按使用范围扫描（有 _MAX_GRID_AREA 上限保护）
    bounds = _used_bounds(sheet)
    if bounds is None:
        return coords
    r0, c0, r1, c1 = bounds
    if (r1 - r0 + 1) * (c1 - c0 + 1) > _MAX_GRID_AREA:
        return coords
    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            value = sheet.cell(row=r + 1, column=c + 1).value
            if value is not None and str(value) != "":
                coords.add((r, c))
    return coords


def _bands_from_sorted(values: list[int]) -> list[tuple[int, int]]:
    """排序值 -> 连续带 [(start, end)]（间隔 >1 处断开）."""
    if not values:
        return []
    bands: list[tuple[int, int]] = []
    start = prev = values[0]
    for v in values[1:]:
        if v - prev > 1:
            bands.append((start, prev))
            start = v
        prev = v
    bands.append((start, prev))
    return bands


def _parse_ref(ref: str) -> tuple[int, int, int, int]:
    """A1:C4 -> (min_col, min_row, max_col, max_row)（1-based 含端点）."""
    from openpyxl.utils import range_boundaries

    return range_boundaries(ref)


def _a1_range(row0: int, col0: int, row1: int, col1: int) -> str:
    return f"{_cell_coord(row0, col0)}:{_cell_coord(row1, col1)}"


def _region_block(
    sheet_formula: Worksheet,
    sheet_values: Worksheet,
    region: dict[str, Any],
    used: tuple[int, int, int, int] | None,
) -> BackendBlock | None:
    """region -> table 块（cell 网格 + 合并 + 公式 + 隐藏态）."""
    row0, col0 = region["row0"], region["col0"]
    row1, col1 = region["row1"], region["col1"]
    rows = row1 - row0 + 1
    cols = col1 - col0 + 1
    grid_clamped = bool(
        used is not None and (
            rows * cols > _MAX_GRID_AREA
            or row1 + 1 > _MAX_GRID_EDGE or col1 + 1 > _MAX_GRID_EDGE
        )
    )

    # 合并区域（裁剪到 region 内）：原点 -> (row_span, col_span)
    origin_spans: dict[tuple[int, int], tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()
    clamped_merges = 0
    for merged in sheet_formula.merged_cells.ranges:
        mr0, mc0 = merged.min_row - 1, merged.min_col - 1
        mr1, mc1 = merged.max_row - 1, merged.max_col - 1
        # 与 region 相交才处理
        if mr0 > row1 or mr1 < row0 or mc0 > col1 or mc1 < col0:
            continue
        area = (mr1 - mr0 + 1) * (mc1 - mc0 + 1)
        if area > _MAX_MERGE_AREA:
            clamped_merges += 1
            continue
        top_left = (mr0 - row0, mc0 - col0)
        origin_spans[top_left] = (mr1 - mr0 + 1, mc1 - mc0 + 1)
        for r in range(mr0, mr1 + 1):
            for c in range(mc0, mc1 + 1):
                if row0 <= r <= row1 and col0 <= c <= col1:
                    covered.add((r - row0, c - col0))

    # 隐藏行/列（region 相对索引，可见性证据不静默）
    hidden_rows = [
        r - row0 for r in range(row0, row1 + 1)
        if (sheet_formula.row_dimensions[r + 1].hidden
            if r + 1 in sheet_formula.row_dimensions else False)
    ]
    hidden_cols = [
        c - col0 for c in range(col0, col1 + 1)
        if (sheet_formula.column_dimensions[_col_letter(c + 1)].hidden
            if _col_letter(c + 1) in sheet_formula.column_dimensions else False)
    ]

    cells: list[dict[str, Any]] = []
    evidence = 0
    empty = True
    for r in range(rows):
        for c in range(cols):
            pos = (r, c)
            if pos in covered and pos not in origin_spans:
                continue
            cell = sheet_formula.cell(row=row0 + r + 1, column=col0 + c + 1)
            formula = _formula_of(cell.value)
            display = sheet_values.cell(
                row=row0 + r + 1, column=col0 + c + 1
            ).value
            if formula is None and display is None and pos not in origin_spans:
                continue
            empty = False
            row_span, col_span = origin_spans.get(pos, (1, 1))
            cells.append({
                "row_index": r,
                "column_index": c,
                "text": "" if display is None else str(display),
                "row_span": row_span,
                "column_span": col_span,
                "is_header": r == 0,  # 首行表头约定（excel_table 亦然）
                "formula": formula,
                "evidence_index": evidence,
            })
            evidence += 1
    if empty:
        return None

    structure: dict[str, Any] = {
        "rows": rows, "cols": cols, "cells": cells,
        "region_origin": region["origin"],
        "region_ref": region["ref"],
        "origin_row": row0, "origin_col": col0,
    }
    if region["header_rows"] == 0:
        structure["header_convention"] = "first_row"  # 约定而非检测事实
    if hidden_rows:
        structure["hidden_rows"] = hidden_rows
    if hidden_cols:
        structure["hidden_columns"] = hidden_cols
    if clamped_merges or grid_clamped:
        structure["clamped_geometry"] = {
            "merges": clamped_merges, "grid": grid_clamped,
        }
    return BackendBlock(
        block_type="table",
        text="",
        container_ref={"container_type": "sheet", "name": sheet_formula.title},
        native_ref={
            "sheet": sheet_formula.title,
            "region": region["ref"],
            "origin": region["origin"],
        },
        structure=structure,
    )


def _formula_of(value: Any) -> str | None:
    """公式判定：openpyxl 中公式格的 value 是 "=..." 字符串."""
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def _col_letter(index_1based: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(index_1based)


# ---------------------------------------------------------------------------
# Normalizer（SRS §C07）
# ---------------------------------------------------------------------------

class XlsxNormalizer(BaseNativeNormalizer):
    """XLSX backend artifact -> Parse IR：workbook/sheet 层级 + cell 证据."""

    normalizer_version = "native-xlsx@2"
    _default_fingerprints = {NATIVE_XLSX_PARSER_ID: NATIVE_XLSX_FINGERPRINT}
    _element_type_map = {"table": "table"}

    def _build_containers(self, artifact) -> tuple[Container, ...]:
        containers = [Container(
            container_id=WORKBOOK_CONTAINER_ID,
            container_type="workbook",
            order_index=0,
            name="workbook",
        )]
        seen: list[str] = []
        for b in artifact.blocks:
            if not b.container_ref:
                continue
            name = b.container_ref.get("name")
            if name is not None and name not in seen:
                seen.append(name)
        for order, name in enumerate(seen):
            containers.append(Container(
                container_id=f"c-sheet-{order}",
                container_type="sheet",
                order_index=order + 1,
                name=name,
                parent_container_id=WORKBOOK_CONTAINER_ID,
            ))
        return tuple(containers)

    def _container_id_for(self, block, containers) -> str | None:
        ref = block.container_ref or {}
        name = ref.get("name")
        for container in containers:
            if container.container_type == "sheet" and container.name == name:
                return container.container_id
        return None

    def _make_spans(self, element_id, block, container_id) -> tuple[EvidenceSpan, ...]:
        """元素级 span：sheet + region 定位."""
        native = dict(block.native_ref) if block.native_ref else {}
        return (EvidenceSpan(
            span_id=f"{element_id}-s0",
            native_ref=native or None,
            raw_text=None,
        ),)

    def _make_cell_spans(
        self, element_id, block, container_id
    ) -> tuple[EvidenceSpan, ...]:
        """每 cell 一个 EvidenceSpan（native_ref=sheet+绝对 A1 坐标）."""
        structure = block.structure or {}
        raw_cells = structure.get("cells") or []
        sheet_name = (block.native_ref or {}).get("sheet")
        origin_row = int(structure.get("origin_row") or 0)
        origin_col = int(structure.get("origin_col") or 0)
        spans: list[EvidenceSpan] = []
        for k, cell in enumerate(raw_cells):
            coord = _cell_coord(
                origin_row + int(cell["row_index"]),
                origin_col + int(cell["column_index"]),
            )
            spans.append(EvidenceSpan(
                span_id=f"{element_id}-cell-{k:04d}",
                native_ref={"sheet": sheet_name, "cell": coord},
                raw_text=str(cell.get("text", "")) or None,
            ))
        return tuple(spans)


def _cell_coord(row_index: int, column_index: int) -> str:
    """0-based (row, col) -> A1 坐标（列字母 + 1-based 行号）."""
    from openpyxl.utils import get_column_letter

    return f"{get_column_letter(column_index + 1)}{row_index + 1}"


__all__ = [
    "NATIVE_XLSX_FINGERPRINT",
    "NATIVE_XLSX_MIMES",
    "NATIVE_XLSX_PARSER_ID",
    "NATIVE_XLSX_VERSION",
    "NativeXlsxParser",
    "WORKBOOK_CONTAINER_ID",
    "XlsxNormalizer",
]
