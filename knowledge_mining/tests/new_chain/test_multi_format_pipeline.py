"""新管线端到端（目标验收层）：多类型文档 → 工程化解析 → 结构化数据 → 对象存储.

对齐业务目标：解析不同类型文档获取结构化数据，资产（源文件 / Parse IR）
落对象存储。全链只走生产组合根 ``build_new_chain_services``（memory 形态，
对象存储用同一 Port 的 Fake 文件实现——与 MinIO 实现共用接口契约）：

  source 对象注册 → DocumentParseFacade（质量门控 + 快照转正）→
  SegmentCompileFacade（rows 档）→ ParseResultReadService（结构化数据视图）。

样例构造复用跨格式契约测试（``test_cross_format_contract``）的七格式
builder，保证本层与 adapter 契约层测试同一份格式输入。
"""
from __future__ import annotations

import hashlib
from datetime import datetime, timezone
from types import SimpleNamespace

import pytest

from knowledge_mining.mining.contracts.file_management import StorageObjectRecord
from knowledge_mining.mining.contracts.storage.types import (
    ObjectLocation,
    PutOptions,
)
from knowledge_mining.mining.infra.object_store.fake import FakeObjectStore
from knowledge_mining.mining.file_management.repositories_memory import (
    MemoryDocumentCurrentContentRepository,
    MemoryStorageObjectRepository,
)
from knowledge_mining.mining.segment_compiler.repositories_memory import (
    MemorySegmentStore,
)
from knowledge_mining.mining.shadow_parse.repositories_memory import (
    MemoryParseAttemptRepository,
    MemoryParseRunRepository,
)
from knowledge_mining.mining.snapshot_store.read_service import (
    ParseResultReadService,
)
from knowledge_mining.mining.snapshot_store.repositories_memory import (
    MemorySnapshotRepository,
)
from knowledge_mining.mining.workflow.new_chain_services import (
    build_new_chain_services,
)
from knowledge_mining.tests.parse_adapters.test_cross_format_contract import (
    _docx_bytes,
    _html_bytes,
    _md_bytes,
    _pdf_with_table,
    _pptx_bytes,
    _txt_bytes,
)

MIME = {
    "md": "text/markdown",
    "txt": "text/plain",
    "docx": "application/vnd.openxmlformats-officedocument."
            "wordprocessingml.document",
    "xlsx": "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet",
    "pptx": "application/vnd.openxmlformats-officedocument."
            "presentationml.presentation",
    "html": "text/html",
    "pdf": "application/pdf",
}
#: 样例内容含表格的格式——表格结构化（tables + 表格行切片）是验收断言项。
TABLE_FORMATS = {"md", "docx", "xlsx", "pptx", "html", "pdf"}
#: 样例内容带显式标题层级的格式——大纲树断言项。
OUTLINE_FORMATS = {"md", "docx", "html", "pptx"}


def _sample_bytes(fmt: str) -> bytes:
    builders = {
        "md": _md_bytes, "txt": _txt_bytes, "docx": _docx_bytes,
        "xlsx": None,  # 由下方内联构造（契约测试未提供独立 builder）
        "pptx": _pptx_bytes, "html": _html_bytes, "pdf": _pdf_with_table,
    }
    builder = builders[fmt]
    if builder is not None:
        return builder()
    from openpyxl import Workbook
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    ws["A1"], ws["B1"] = "表头A", "表头B"
    ws["A2"], ws["B2"] = "a1", "b1"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _chunked(payload: bytes):
    yield payload


async def _seed_document(
    store, objects, documents, *, fmt: str, data: bytes,
    doc_id: str | None = None, document_key: str | None = None,
):
    """M1 式建档：内容寻址对象上传 + 注册行 + 文档当前内容指针."""
    sha = hashlib.sha256(data).hexdigest()
    bucket = "e2e-source"
    key = f"v1/{sha[:2]}/{sha}"
    await store.put_stream(
        ObjectLocation(bucket=bucket, object_key=key), _chunked(data),
        PutOptions(artifact_class="source", expected_sha256=sha),
    )
    so_id = f"so_{fmt}"
    await objects.register(StorageObjectRecord(
        id=so_id, provider="fake", bucket=bucket, object_key=key,
        object_version_id=None, sha256=sha, size=len(data),
        mime=MIME[fmt], artifact_class="source", state="AVAILABLE",
        created_at=datetime.now(timezone.utc).isoformat(),
    ))
    doc_id = doc_id or f"doc-e2e-{fmt}"
    document_key = document_key or f"sample.{fmt}"
    await documents.create_document(
        kb_id="kb-e2e", document_id=doc_id, folder_id=None, owner_id=None,
        document_name=document_key, document_type="other",
        storage_object_id=so_id, source_raw_hash=sha,
    )
    return SimpleNamespace(
        document_id=doc_id, document_key=document_key,
        file_type=fmt, mime=MIME[fmt],
    )


@pytest.mark.parametrize("fmt", sorted(MIME))
@pytest.mark.asyncio
async def test_multi_format_document_to_structured_data(fmt: str, tmp_path):
    """一种格式一条全链：解析出结构化数据，切片与 IR 全部落对象存储."""
    store = FakeObjectStore(str(tmp_path / "objects"))
    objects = MemoryStorageObjectRepository()
    documents = MemoryDocumentCurrentContentRepository()
    parse_runs = MemoryParseRunRepository()
    attempts = MemoryParseAttemptRepository()
    snapshots = MemorySnapshotRepository()
    segment_store = MemorySegmentStore()
    services = build_new_chain_services(
        bucket_prefix="e2e-", object_store=store, storage_objects=objects,
        documents=documents, parse_runs=parse_runs, attempts=attempts,
        snapshots=snapshots, segment_store=segment_store,
    )

    data = _sample_bytes(fmt)
    raw = await _seed_document(
        store, objects, documents, fmt=fmt, data=data,
    )

    # 1. 解析：质量门控（PASS/WARN 才有快照）+ Parse IR 落对象存储
    outcome = services.document_parse_service.parse_document(
        raw, params={}, domain="e2e", run_document_id=f"rd-{fmt}",
    )
    assert outcome is not None, f"{fmt}: document skipped"
    assert outcome.status == "SUCCEEDED", f"{fmt}: parse {outcome.status}"
    assert outcome.snapshot_id, f"{fmt}: no snapshot committed"
    snapshot = await snapshots.get(outcome.snapshot_id)
    assert snapshot is not None and snapshot.quality_status in ("PASS", "WARN"), (
        f"{fmt}: quality gate rejected: "
        f"{getattr(snapshot, 'quality_status', None)}"
    )
    assert outcome.parse_ir_storage_object_id, f"{fmt}: Parse IR not stored"

    # 2. 切片编译：rows 档——表格行独立成段
    compiled = services.segment_compile_service.compile_for_snapshot(
        snapshot_id=outcome.snapshot_id,
        parse_ir_storage_object_id=outcome.parse_ir_storage_object_id,
        params={"tableView": "rows"},
    )
    assert compiled.segment_count >= 1, f"{fmt}: no segments compiled"

    # 3. 结构化数据视图（前端「结构化数据」页签同源）
    reader = ParseResultReadService(
        snapshots=snapshots, storage_objects=objects, object_store=store,
        segment_store=segment_store, documents=documents,
    )
    result = await reader.get_parse_result(domain="e2e", document_id=raw.document_id)
    assert result is not None, f"{fmt}: structured data view missing"

    outline = result.get("outline") or []
    tables = result.get("tables") or []
    segments = (result.get("segments") or {}).get("items") or []
    snapshot = result.get("snapshot") or {}

    assert snapshot.get("quality_status") in ("PASS", "WARN")
    if fmt in OUTLINE_FORMATS:
        assert outline, f"{fmt}: outline tree missing"
    if fmt in TABLE_FORMATS:
        assert tables, f"{fmt}: table grid missing from structured data"
        assert any(t.get("rows") for t in tables), f"{fmt}: table has no rows"
        table_segs = [s for s in segments if s.get("block_type") == "table"]
        assert table_segs, f"{fmt}: no table segments (rows view)"
    assert segments, f"{fmt}: segments missing from structured data"
    # 证据链：每条切片都能指回解析元素（element_ids 非空）
    assert all(s.get("element_ids") for s in segments), (
        f"{fmt}: segments lost element evidence links"
    )


@pytest.mark.asyncio
async def test_same_content_shares_snapshot_and_compiles_once(tmp_path):
    """同内容多文档：共享一个快照，重复编译幂等（不撞唯一键）."""
    store = FakeObjectStore(str(tmp_path / "objects"))
    objects = MemoryStorageObjectRepository()
    documents = MemoryDocumentCurrentContentRepository()
    snapshots = MemorySnapshotRepository()
    services = build_new_chain_services(
        bucket_prefix="e2e-", object_store=store, storage_objects=objects,
        documents=documents, snapshots=snapshots,
        segment_store=MemorySegmentStore(),
    )

    data = _md_bytes()
    raws = []
    for suffix in ("a", "b"):
        raws.append(await _seed_document(
            store, objects, documents, fmt="md", data=data,
            doc_id=f"doc-dup-{suffix}", document_key=f"duplicate-{suffix}.md",
        ))

    outcomes = [
        services.document_parse_service.parse_document(
            raw, params={}, domain="e2e", run_document_id=f"rd-dup-{i}",
        )
        for i, raw in enumerate(raws)
    ]
    assert outcomes[0].snapshot_id == outcomes[1].snapshot_id

    first = services.segment_compile_service.compile_for_snapshot(
        snapshot_id=outcomes[0].snapshot_id,
        parse_ir_storage_object_id=outcomes[0].parse_ir_storage_object_id,
        params={"tableView": "rows"},
    )
    second = services.segment_compile_service.compile_for_snapshot(
        snapshot_id=outcomes[1].snapshot_id,
        parse_ir_storage_object_id=outcomes[1].parse_ir_storage_object_id,
        params={"tableView": "rows"},
    )
    assert second.segment_count == first.segment_count
    assert second.compiler_fingerprint == first.compiler_fingerprint
