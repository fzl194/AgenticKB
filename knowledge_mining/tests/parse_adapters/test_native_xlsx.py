"""``NativeXlsxParser`` + ``XlsxNormalizer`` 单元测试（M3, SRS §C06/§C07）.

fixture 用 openpyxl 在测试内生成 bytes。覆盖：workbook→sheet 容器层级、
双 sheet、合并区域 span、公式与展示值分离（data_only 双读模式）、
cell 级 native_ref 证据（TableCell.source_span_id）、TableAsset 网格、
IR 断言链、错误归一。
"""
from __future__ import annotations

import io

import pytest

from knowledge_mining.mining.contracts.parse_ir import validate
from knowledge_mining.mining.contracts.parser_adapter import (
    ParserAdapterError,
    UnsupportedFormat,
)
from knowledge_mining.mining.parse_adapters.native.native_xlsx import (
    NATIVE_XLSX_FINGERPRINT,
    NativeXlsxParser,
    XlsxNormalizer,
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
RAW_HASH = "ef" * 32


def _build_xlsx_bytes() -> bytes:
    """双 sheet：Summary（字面值 + 公式）、Merged（2x2 合并区域）."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "名称"
    ws1["B1"] = "数量"
    ws1["A2"] = "苹果"
    ws1["B2"] = 42
    ws1["A3"] = "合计"
    ws1["B3"] = "=B2*2"

    ws2 = wb.create_sheet("Merged")
    ws2["A1"] = "锚点"
    ws2.merge_cells("A1:B2")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def xlsx_bytes() -> bytes:
    return _build_xlsx_bytes()


@pytest.fixture
def parser() -> NativeXlsxParser:
    return NativeXlsxParser()


@pytest.fixture
def normalizer() -> XlsxNormalizer:
    return XlsxNormalizer()


def _normalize(parser, normalizer, data: bytes):
    artifact = parser.parse(data, mime=XLSX_MIME)
    return normalizer.normalize(artifact, source_raw_hash=RAW_HASH)


def _sheet_table(doc, sheet_name: str):
    """按容器名取该 sheet 的 table element + TableAsset."""
    for el in doc.elements:
        if el.element_type != "table":
            continue
        span = el.source_spans[0]
        if span.native_ref and span.native_ref.get("sheet") == sheet_name:
            asset = doc.structured_assets[f"{el.element_id}-table"]
            return el, asset
    raise AssertionError(f"no table element for sheet {sheet_name!r}")


# ---------------------------------------------------------------------------
# RED 1: parse 层
# ---------------------------------------------------------------------------

def test_parse_two_sheet_blocks_with_container_refs(parser, xlsx_bytes) -> None:
    artifact = parser.parse(xlsx_bytes, mime=XLSX_MIME)

    assert artifact.parser_id == "native_xlsx"
    assert [b.block_type for b in artifact.blocks] == ["table", "table"]
    assert [b.container_ref["name"] for b in artifact.blocks] == [
        "Summary", "Merged",
    ]
    summary = artifact.blocks[0]
    assert summary.structure["rows"] == 3
    assert summary.structure["cols"] == 2


def test_parse_unsupported_mime_and_corrupt(parser) -> None:
    with pytest.raises(UnsupportedFormat):
        parser.parse(b"x", mime="text/csv")
    with pytest.raises(ParserAdapterError):
        parser.parse(b"garbage-not-zip", mime=XLSX_MIME)


# ---------------------------------------------------------------------------
# RED 2: IR 断言链 —— 容器层级 / 网格 / 公式 / 证据
# ---------------------------------------------------------------------------

def test_container_hierarchy_workbook_to_sheets(parser, normalizer, xlsx_bytes) -> None:
    doc = _normalize(parser, normalizer, xlsx_bytes)
    assert validate(doc).valid

    assert [c.container_type for c in doc.containers] == [
        "workbook", "sheet", "sheet",
    ]
    workbook, sheet1, sheet2 = doc.containers
    assert workbook.parent_container_id is None
    assert sheet1.parent_container_id == workbook.container_id
    assert sheet2.parent_container_id == workbook.container_id
    assert (sheet1.name, sheet2.name) == ("Summary", "Merged")
    assert [c.order_index for c in doc.containers] == [0, 1, 2]


def test_summary_grid_and_formula_display_separation(
    parser, normalizer, xlsx_bytes
) -> None:
    doc = _normalize(parser, normalizer, xlsx_bytes)
    assert validate(doc).valid

    _, asset = _sheet_table(doc, "Summary")
    assert asset.rows == 3
    assert asset.columns == 2
    grid = {(c.row_index, c.column_index): c for c in asset.cells}

    # 字面值：文本与库读出的展示值一致，无公式
    assert grid[(1, 1)].text == "42"
    assert grid[(1, 1)].formula is None
    # 公式单元格：公式与展示值分离；openpyxl 写出的文件无缓存值，
    # 展示值缺失 -> text 为空串（不得伪造缓存结果）
    assert grid[(2, 1)].formula == "=B2*2"
    assert grid[(2, 1)].text == ""
    # 首行 is_header 约定
    assert grid[(0, 0)].is_header and grid[(0, 1)].is_header
    assert not grid[(1, 0)].is_header
    assert asset.header_regions == ((0, 0),)


def test_merged_region_expands_to_span(parser, normalizer, xlsx_bytes) -> None:
    doc = _normalize(parser, normalizer, xlsx_bytes)

    _, asset = _sheet_table(doc, "Merged")
    assert asset.rows == 2
    assert asset.columns == 2
    grid = {(c.row_index, c.column_index): c for c in asset.cells}
    assert list(grid) == [(0, 0)]  # 仅合并原点产 cell
    assert grid[(0, 0)].text == "锚点"
    assert grid[(0, 0)].row_span == 2
    assert grid[(0, 0)].column_span == 2


def test_cell_native_ref_evidence_links(parser, normalizer, xlsx_bytes) -> None:
    doc = _normalize(parser, normalizer, xlsx_bytes)

    el, asset = _sheet_table(doc, "Summary")
    span_by_cell = {
        s.native_ref["cell"]: s for s in el.source_spans
        if s.native_ref is not None
    }
    assert span_by_cell["B3"].native_ref == {"sheet": "Summary", "cell": "B3"}
    # TableCell.source_span_id -> 元素上声明的 EvidenceSpan
    for cell in asset.cells:
        assert cell.source_span_id is not None
        declared = {s.span_id for s in el.source_spans}
        assert cell.source_span_id in declared
    b3 = next(c for c in asset.cells if (c.row_index, c.column_index) == (2, 1))
    assert b3.source_span_id == span_by_cell["B3"].span_id


def test_ir_chain_relations_and_fingerprint(parser, normalizer, xlsx_bytes) -> None:
    doc = _normalize(parser, normalizer, xlsx_bytes)
    assert validate(doc).valid

    # 两个 sheet table element 按阅读顺序链接
    reading = [
        r for r in doc.relations if r.relation_type == "next_in_reading_order"
    ]
    assert len(reading) == 1
    # 各 element 挂在对应 sheet 容器上
    tables = [e for e in doc.elements if e.element_type == "table"]
    assert all(len(e.page_span_ids) == 1 for e in tables)
    sheet_ids = {c.name: c.container_id for c in doc.containers}
    summary_el, _ = _sheet_table(doc, "Summary")
    merged_el, _ = _sheet_table(doc, "Merged")
    assert summary_el.page_span_ids == (sheet_ids["Summary"],)
    assert merged_el.page_span_ids == (sheet_ids["Merged"],)

    assert doc.source_identity.parser_fingerprint == NATIVE_XLSX_FINGERPRINT
    assert doc.source_identity.parser_fingerprint == (
        "native_xlsx@1.0.0#openpyxl-3.1.5"
    )


def test_malicious_sparse_grid_clamped() -> None:
    """评审 HIGH-2 回归：稀疏远端格撑大网格维度 -> 截断到 10k + 可见标记.

    注：mergeCell A1:XFD1048576 形态会让 openpyxl **打开期**本身挂死
    （库内部展开合并区域，发生在适配层截断逻辑之前）——该形态属
    库级已知边界（记录于 M3 报告已知缺口），依赖上游 M1 intake 的
    文件级限制兜底；本测试覆盖适配层能防的"能打开但网格巨大"形态。
    """
    import io
    import time
    from openpyxl import Workbook
    from knowledge_mining.mining.parse_adapters.native.native_xlsx import (
        NativeXlsxParser,
    )

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "head"
    ws.cell(row=1, column=16384, value="far")  # 撑大 max_column 至极限
    buf = io.BytesIO(); wb.save(buf)
    parser = NativeXlsxParser()
    start = time.monotonic()
    artifact = parser.parse(
        buf.getvalue(),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    elapsed = time.monotonic() - start
    assert elapsed < 5.0, f"parse took {elapsed:.2f}s (DoS surface open)"
    tables = [b for b in artifact.blocks if b.block_type == "table"]
    assert tables, "sheet lost entirely"
    assert tables[0].structure["rows"] <= 10_000
    assert tables[0].structure["cols"] <= 10_000
    assert tables[0].structure.get("clamped_geometry", {}).get("grid") is True
