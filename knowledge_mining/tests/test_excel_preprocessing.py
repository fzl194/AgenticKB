from openpyxl import Workbook
import pytest

from knowledge_mining.mining.infra.excel_config import ExcelConfig
from knowledge_mining.mining.ingestion.errors import PreprocessingError
from knowledge_mining.mining.ingestion.excel_preprocessing import excel_to_markdown


def _config(**overrides):
    values = {
        "max_sheets": 10,
        "max_nonempty_cells": 10_000,
        "table_chunk_target_tokens": 420,
    }
    values.update(overrides)
    return ExcelConfig(**values)


def test_xlsx_becomes_sheet_scoped_markdown(tmp_path):
    path = tmp_path / "设备台账.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "核心网设备"
    ws.append(["设备名称", "厂家", "状态"])
    ws.append(["AMF01", "华为", "运行中"])
    ws.append(["SMF01", "中兴", "运行中"])
    wb.save(path)

    result = excel_to_markdown(path, _config())

    assert result.status == "success"
    assert "# 设备台账.xlsx" in result.markdown
    assert "## 工作表：核心网设备" in result.markdown
    assert "表格 A1:C3" in result.markdown
    assert "| 设备名称 | 厂家 | 状态 |" in result.markdown
    assert result.summary["table_region_count"] == 1


def test_hidden_sheet_is_marked_and_empty_sheet_is_counted(tmp_path):
    path = tmp_path / "book.xlsx"
    wb = Workbook()
    wb.active.append(["A", "B"])
    hidden = wb.create_sheet("隐藏")
    hidden.sheet_state = "hidden"
    hidden.append(["参数", "值"])
    wb.create_sheet("空表")
    wb.save(path)

    result = excel_to_markdown(path)

    assert "## 工作表：隐藏（隐藏）" in result.markdown
    assert result.summary["skipped_empty_sheet_count"] == 1


def test_workbook_limit_fails_without_silent_truncation(tmp_path):
    path = tmp_path / "large.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append(["1", "2"])
    wb.save(path)

    with pytest.raises(PreprocessingError) as caught:
        excel_to_markdown(path, _config(max_nonempty_cells=2))

    assert caught.value.code == "excel_limits_exceeded"


def test_chunking_is_deterministic_repeats_headers_and_keeps_every_row(tmp_path):
    path = tmp_path / "many.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["设备名称", "状态"])
    for row_number in range(100):
        ws.append([f"设备-{row_number}", "正常"])
    wb.save(path)
    config = _config(table_chunk_target_tokens=30)

    first = excel_to_markdown(path, config)
    second = excel_to_markdown(path, config)

    assert first.markdown == second.markdown
    assert first.markdown.count("| 设备名称 | 状态 |") > 1
    for row_number in range(100):
        assert first.markdown.count(f"设备-{row_number} |") == 1


def test_markdown_escapes_table_delimiters_backslashes_and_newlines(tmp_path):
    path = tmp_path / "escape.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["名称", "说明"])
    ws.append(["A|B", "路径\\节点\n下一行"])
    ws.append(["C", "正常"])
    wb.save(path)

    result = excel_to_markdown(path, _config())

    assert "A\\|B" in result.markdown
    assert "路径\\\\节点<br>下一行" in result.markdown


def test_ingest_xlsx_routes_markdown_and_preserves_source_metadata(tmp_path):
    path = tmp_path / "inventory.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["name", "status"])
    ws.append(["AMF", "active"])
    wb.save(path)

    from knowledge_mining.mining.ingestion import get_mime_type, ingest_directory

    docs, summary = ingest_directory(tmp_path)

    assert len(docs) == 1
    assert docs[0].file_type == "markdown"
    assert docs[0].source_uri == str(path)
    assert docs[0].metadata_json["source_format"] == "xlsx"
    assert docs[0].metadata_json["preprocess_status"] == "success"
    assert docs[0].metadata_json["excel_summary"]["table_region_count"] == 1
    assert (
        get_mime_type("xlsx")
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert summary["preprocessed_excels"] == 1
