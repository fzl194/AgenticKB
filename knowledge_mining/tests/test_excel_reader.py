from datetime import datetime

from openpyxl import Workbook
import pytest
import xlwt

from knowledge_mining.mining.ingestion.errors import PreprocessingError
from knowledge_mining.mining.ingestion.excel_reader import read_excel_workbook


def test_read_xlsx_preserves_sheet_state_merge_and_values(tmp_path):
    path = tmp_path / "设备.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "设备"
    ws.merge_cells("A1:B1")
    ws["A1"] = "基本信息"
    ws.append(["名称", "上线日期"])
    ws.append(["AMF01", datetime(2026, 8, 7)])
    hidden = wb.create_sheet("隐藏参数")
    hidden.sheet_state = "hidden"
    hidden.append(["参数", "值"])
    hidden.append(["超时", 30])
    wb.save(path)

    workbook = read_excel_workbook(path)

    assert [sheet.name for sheet in workbook.sheets] == ["设备", "隐藏参数"]
    assert workbook.sheets[1].hidden is True
    assert workbook.sheets[0].rows[0] == ("基本信息", "基本信息")
    assert workbook.sheets[0].rows[2][1] == "2026-08-07T00:00:00"


def test_read_xls_normalizes_values_and_merge(tmp_path):
    path = tmp_path / "legacy.xls"
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Sheet1")
    ws.write_merge(0, 0, 0, 1, "基本信息")
    ws.write(1, 0, "名称")
    ws.write(1, 1, "数量")
    ws.write(2, 0, "SMF")
    ws.write(2, 1, 2)
    wb.save(str(path))

    workbook = read_excel_workbook(path)

    assert workbook.source_format == "xls"
    assert workbook.sheets[0].rows == (
        ("基本信息", "基本信息"),
        ("名称", "数量"),
        ("SMF", "2"),
    )


def test_read_xlsx_normalizes_special_values_and_reports_warnings(tmp_path):
    path = tmp_path / "values.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = True
    ws["B1"] = 0.125
    ws["B1"].number_format = "0.0%"
    ws["C1"] = "=1+1"
    ws["D1"] = "#DIV/0!"
    ws["D1"].data_type = "e"
    wb.save(path)

    workbook = read_excel_workbook(path)

    assert workbook.sheets[0].rows[0] == ("true", "12.5%", "", "#DIV/0!")
    warning_codes = {warning.code for warning in workbook.sheets[0].formula_warnings}
    assert warning_codes == {"excel_formula_cache_missing", "excel_cell_error"}


def test_read_excel_rejects_unknown_suffix(tmp_path):
    path = tmp_path / "values.csv"
    path.write_text("a,b", encoding="utf-8")

    with pytest.raises(PreprocessingError) as caught:
        read_excel_workbook(path)

    assert caught.value.code == "excel_unsupported_format"
